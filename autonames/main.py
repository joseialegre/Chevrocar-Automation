from colorama import Fore, Style
from openpyxl import load_workbook
import csv
from openpyxl.styles import Alignment

archivo1 = input("ingrese nombre de archivo de Publicaciones:     ")
archivo2 = input("ingrese nombre de archivo de TiendaNube:     ")

archivo1= archivo1+".xlsx"
archivo2= archivo2+".csv"

wb1 = load_workbook(archivo1)



sheet1 = wb1['Publicaciones']

azul = Fore.BLUE
rojo = Fore.RED
verde = Fore.GREEN
amarillo = Fore.YELLOW
reset = Style.RESET_ALL


with open(archivo2, 'r') as csvfile:

    reader = csv.reader(csvfile, delimiter=';')
    lista_tiendanube = list(reader)

    for filaPublicaciones in sheet1.iter_rows(min_row=5, values_only=True):
        for index, filaTienda in enumerate(lista_tiendanube):

            if str(filaTienda[16]) == str(filaPublicaciones[2]):
                print(amarillo + "Copiando los valores de... " + azul + str(filaTienda[16]) + reset)
                print("Antes: " + rojo + str(filaTienda[1]) + reset + " -->" + " Después: " + azul + verde + str(filaPublicaciones[3]) + reset)
                lista_tiendanube[index][1] = filaPublicaciones[3]


with open(archivo2, 'w', newline='') as csvfile:
    writer = csv.writer(csvfile, delimiter=';')
    writer.writerows(lista_tiendanube)



wb1.close()


# SKU = row1 archivo publicacion columna A = 0
# SKU = row2 archivo tiendanube columna Q = 16

# PRECIO para Tienda columna J (numero 9) es el precio
# PRECIO de Exportar columna C (numero 2)

# NOMBRE en Tienda columna B
# NOMBRE en Publicaciones columna D
