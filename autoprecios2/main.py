from openpyxl import load_workbook, Workbook
from colorama import Fore, Style
import sys

import csv
from openpyxl.styles import Alignment

archivo1 = input("ingrese nombre de archivo de Exportar:     ")
archivo2 = input("ingrese nombre de archivo de TiendaNube:     ")

archivo2= archivo2+".csv"

wb1 = load_workbook(archivo1+".xlsx")
sheet1 = wb1.active

new_sheet = wb1.create_sheet('Temp')

wb3 = Workbook()
sheet3 =wb3.active
sheet3['A1'] = "SKU"
sheet3['B1'] = "Precio Acutal"
sheet3['C1'] = "Precio Anterior"
sheet3['D1'] = "Diferencia"
sheet3['E1'] = "Variacion"


#########################
azul = Fore.BLUE
rojo = Fore.RED
verde = Fore.GREEN
amarillo = Fore.YELLOW
reset = Style.RESET_ALL
#########################

# ELIMINO LAS COLUMNAS CON FALSE, CREO OTRO SHEET
# LUEGO REEMPLAZO EL ORIGINAL
for row in sheet1.iter_rows(min_row=1, max_row=sheet1.max_row, values_only=True):
    if row[5] != False:
        new_sheet.append(row)
        print(verde + row[1] + " OK")
    else:
        print(rojo+row[1]+"ELIMINADO")
wb1.remove(sheet1)
new_sheet.title = sheet1.title
wb1.save(archivo1+".xlsx")

with open(archivo2, 'r') as csvfile:
    reader = csv.reader(csvfile, delimiter=';')
    lista_tiendanube = list(reader)

    for filaExportar in new_sheet.iter_rows(min_row=2, values_only=True):
        for index, filaTienda in enumerate(lista_tiendanube):
            # Comparamos SKU
            if str(filaTienda[16]) == str(filaExportar[1]):
                print(amarillo + "Copiando los valores de... " +azul+ str(filaTienda[16]) + reset)
                print("Antes: " + rojo + str(filaTienda[9]) + reset + " -->" +  " Después: " + azul + verde + str(filaExportar[22]) + reset)
                sheet3[f'A{index + 1}'] = filaExportar[1]
                sheet3[f'B{index + 1}'] = filaExportar[22]
                sheet3[f'C{index + 1}'] = filaTienda[9]
                sheet3[f'D{index + 1}'] = float(filaExportar[22])-float(filaTienda[9])
                if float(filaTienda[9]) > float(filaExportar[22]):
                    sheet3[f'E{index + 1}'] = "Disminución"
                if float(filaTienda[9]) < float(filaExportar[22]):
                    sheet3[f'E{index + 1}'] = "Aumento"
                if float(filaTienda[9]) == float(filaExportar[22]):
                    sheet3[f'E{index + 1}'] = "Sin Cambio"

                lista_tiendanube[index][9] = filaExportar[22]

with open(archivo2, 'w', newline='') as csvfile:
    writer = csv.writer(csvfile, delimiter=';')
    writer.writerows(lista_tiendanube)

for row in sheet3.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

print("Goodbye :)")
wb1.close()
wb3.save("Precios.xlsx")


# SKU de Exportar columna B = 1
# SKU de tiendanube columna Q = 16

# PRECIO para Tienda columna J (numero 9) es el precio
# PRECIO de Exportar columna W (numero 22)